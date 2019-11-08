from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
import os
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill,NamedStyle
os.chdir( r'F:\Personal\Excel') #set working directory

# 初始化配置 && pandas对象全部显示
filename = 'SalesB.xlsx'
wb = load_workbook(filename)
pd.set_option('display.max_columns', 1000)
pd.set_option('display.width', 1000)
pd.set_option('display.max_colwidth', 1000)

# 生成pandas对象
def get_df(fileName,sheet_name = 'Sales'):
    """
    生成pandas 对象
    :param fileName:原始excel名称
    :return:pandas对象
    """
    excel_reader = pd.ExcelFile(fileName)


    # df1为原始表
    df1 = pd.read_excel(excel_reader, sheet_name)


    # df2为汇总季度销售额并且排序表
    df2 = pd.read_excel(excel_reader, sheet_name)
    df2['1季度销售额'] = df2.sum(axis=1)                              # 季度销售额为前三个月之和 axis = 1 默认去除非数字列
    if('华北区' in df2['所属部门'].unique()):
        sort_list = ['华北区','华东区','华南区']
    else:
        sort_list = ['华东区','华南区']
    df2['所属部门'] = df2['所属部门'].astype('category')                   # 设置成“category”数据类型
    df2['所属部门'].cat.reorder_categories(sort_list, inplace=True)       # inplace = True，使 recorder_categories生效
    df2 = df2.sort_values(by=["所属部门","1季度销售额"],ascending=(True,False)) # 根据所属部门和季度销售额进行排序
    
    # df3为地区季度销售额汇总
    df3 = df2.groupby(by='所属部门').agg({'1季度销售额':sum}).reset_index()
    return [df1,df2,df3]

def generate_excel(df_list, excel_name='./RESULT.xlsx', sheet_list=['Sales','Report1','Report2']):
    """
    生成excel
    :param sheet_name:
    :param excel_name:
    :param df:
    :return:
    """
    # 如果文件存在，直接添加sheet；不存在则创建新的文件
    excel_writer = pd.ExcelWriter(excel_name, engine='openpyxl')
    # if not os.path.exists(excel_name):
    for i in range(0, len(df_list)):    #遍历将数据写入excel
        df_list[i].to_excel(excel_writer, sheet_name=sheet_list[i],index=None)
    # else:
    #     book = load_workbook(excel_writer.path)
    #     excel_writer.book = book
    #     for i in range(0, len(df_list)):    #遍历将数据写入excel
    #         df_list[i].to_excel(excel_writer, sheet_name=sheet_list[i],index=None)
    excel_writer.save()
    excel_writer.close()

# 设置excel样式
def set_excel_style( excel_name='./Report.xlsx', sheet_list=['Sales','Report1','Report2']):
    book = load_workbook(excel_name)
    for sheet_name in sheet_list:
        sheet = book[sheet_name]

        # 初始化字体样式
        font_not_bold = Font(name='华文中宋',
                            size=11,
                            italic=False,
                            vertAlign=None,
                            underline='none',
                            strike=False,
                            color='FF000000',
                            outline='None')
        # 设置整体的字体
        for row in sheet.rows:
            for cell in row:
                cell.font = font_not_bold

        # 初始化背景颜色样式
        header = NamedStyle(name="header")      #头部效果
        header.fill = PatternFill("solid", fgColor="FFFF00")
        
        east = NamedStyle(name="east")      #头部效果
        east.fill = PatternFill("solid", fgColor="FFFF00")
        north = NamedStyle(name="north")      #头部效果
        north.fill = PatternFill("solid", fgColor="FFFF00")
        south = NamedStyle(name="south")      #头部效果
        south.fill = PatternFill("solid", fgColor="FFFF00")

        #设置背景颜色
        # for row in sheet.rows:
        #     if(row == 0 ):
        #         line = sheet.row_dimensions[row]
        #         line.style = header
        #     key_colomn = 1
        #     if(sheet_name == 'Report2'):
        #         key_colomn = 0
        #     if(row[key_colomn].value == '华南区'):
        #         line = ws.row_dimensions[row]
        #         line.style = header
        # 保存excel
        book.save(excel_name)
        

if __name__ == '__main__':
    df_list = get_df(filename)
    generate_excel(df_list)
    set_excel_style()



